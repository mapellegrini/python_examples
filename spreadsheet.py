#!/usr/bin/python

import openpyxl

#takes a number, starting at 0
#returns the letter used for that row 
#acts as a wrapper for built-in utility function 
def num2rowchar(val):
  return openpyxl.utils.cell.get_column_letter(val+1)

#create a new workbook 
wb = openpyxl.Workbook()

#by default, wb comes with a sheet called "Sheet". Let's delete it
del wb["Sheet"]

#create new worksheets 
ws1 = wb.create_sheet("Sheet1")
ws2 = wb.create_sheet("Sheet2")

#set values 
ws1["A1"]=1
ws1["A2"]=2
ws1["A3"]=3
ws1["B1"]=7
ws1["B2"]=8
ws1["B3"]=9


#set formulas
ws1["C1"]="=SUM(A1:B3)"
ws1["C2"]="=MIN(A1:B3)"
ws1["C3"]="=MAX(A1:B3)"
ws1["C4"]="=STDEV(A1:B3)"

#add color
#rule = openpyxl.formatting.rule.ColorScaleRule(start_type='percentile', start_value=10, start_color='0000FF00', mid_type='percentile', mid_value=50, mid_color='00FFA500', end_type='percentile', end_value=90, end_color='00FF0000')
#tsheet.conditional_formatting.add("B2:" + sschar(col) +"25", rule)

#add chart 
'''
maxrow=len(wb.sheetnames)-len(startsheets) + 1
dowxaxis=openpyxl.chart.Reference(ssheet, min_col=3, min_row=2, max_row=maxrow)
hourxaxis=openpyxl.chart.Reference(ssheet, min_col=4, min_row=2, max_row=maxrow)
avgavg=openpyxl.chart.Reference(ssheet, min_col=5, min_row=2, max_row=maxrow)
avgstdev=openpyxl.chart.Reference(ssheet, min_col=6, min_row=2, max_row=maxrow)

chart1 = openpyxl.chart.ScatterChart(scatterStyle='smoothMarker')
chart1.title = startsheets[1]
chart1.x_axis.title = 'DOW'
chart1.y_axis.title = 'Avgavg Count'
chart1.legend=None
s1 = openpyxl.chart.Series(avgavg, dowxaxis, title=startsheets[1], title_from_data=False)
sp_color = openpyxl.drawing.colors.ColorChoice(prstClr="black")
sp_sppr = openpyxl.chart.shapes.GraphicalProperties(solidFill=sp_color)
s1.marker.spPr.ln.noFill = True
s1.spPr.ln.noFill = True
s1.marker = openpyxl.chart.marker.Marker(symbol=('square'), size=6, spPr=sp_sppr)
chart1.series.append(s1)
wsheet1.add_chart(chart1, "A1")
'''

#save to file 
wb.save("/tmp/openpyxl.xlsx")


#load file
newwb=openpyxl.load_workbook("/tmp/openpyxl.xlsx")
newws2=wb["Sheet2"]
newws2["A1"]="Hello"
newws2["A2"]="I'm"
newws2["A3"]="Cat"

newwb.save("/tmp/openpyxl.xlsx")
